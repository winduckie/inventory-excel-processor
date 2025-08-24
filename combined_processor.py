#!/usr/bin/env python3
"""
Combined Excel Processor with Global Usage Integration

A processor that combines delivery and inventory data from multiple Excel files,
plus global usage data for enhanced inventory projections.
"""

import pandas as pd
import numpy as np
import json
from typing import Dict, List, Tuple, Optional
import logging
import os
from datetime import datetime
import openpyxl

# Set up logging
os.makedirs('logs', exist_ok=True)

# Create a unique log filename with timestamp
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
log_filename = f'logs/combined_processor_{timestamp}.log'

# Set up logging to both file and console
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, mode='w'),
        logging.StreamHandler()  # This keeps console output
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"Logging to file: {log_filename}")

class CombinedProcessor:
    """
    A class to process and combine delivery, inventory, and global usage data.
    """
    
    def __init__(self, delivery_file: str, inventory_file: str, global_usage_file: str = None):
        """
        Initialize the Combined processor.
        
        Args:
            delivery_file (str): Path to the delivery Excel file
            inventory_file (str): Path to the inventory Excel file
            global_usage_file (str, optional): Path to the global usage CSV file
        """
        self.delivery_file = delivery_file
        self.inventory_file = inventory_file
        self.global_usage_file = global_usage_file
        self.delivery_data = {}
        self.inventory_data = {}
        self.global_usage_data = {}
        self.ingredient_mapping = {}
        self.column_mapping = self._load_column_mapping()
        
        # Extract date from input folder for organized output
        self.input_date = self._extract_date_from_input_path()
        self.output_dir = f"processed_data/{self.input_date}" if self.input_date else "processed_data"
        
    def _extract_date_from_input_path(self) -> str:
        """
        Extract date from input folder path (e.g., input/20250811 -> 20250811)
        
        Returns:
            str: Date string or empty string if no date found
        """
        import os
        import re
        
        # Get the directory containing the delivery file
        delivery_dir = os.path.dirname(self.delivery_file)
        
        # Look for date pattern in the path (YYYYMMDD)
        date_pattern = r'(\d{8})'
        match = re.search(date_pattern, delivery_dir)
        
        if match:
            date_str = match.group(1)
            logger.info(f"Detected date from input path: {date_str}")
            return date_str
        
        # If no date found, try to get current date
        current_date = datetime.now().strftime('%Y%m%d')
        logger.info(f"No date found in input path, using current date: {current_date}")
        return current_date
        
    def _load_column_mapping(self) -> Dict:
        """
        Load the column mapping from JSON file.
        
        Returns:
            Dict: Column mapping dictionary
        """
        try:
            with open('config/table_column_mapping.json', 'r') as f:
                mapping = json.load(f)
                return mapping['table_column_mapping']
        except FileNotFoundError:
            logger.error("config/table_column_mapping.json not found")
            return {}
        except Exception as e:
            logger.error(f"Error loading column mapping: {e}")
            return {}
    
    def _load_product_categories(self) -> Dict[str, str]:
        """
        Load product categories from CSV file.
        
        Returns:
            Dict[str, str]: Product to category mapping
        """
        try:
            import csv
            categories = {}
            with open('config/product_categories.csv', 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    product = row['Product'].strip()
                    category = row['Category'].strip()
                    categories[product] = category
            logger.info(f"Loaded {len(categories)} product categories")
            return categories
        except FileNotFoundError:
            logger.warning("config/product_categories.csv not found, creating new file")
            return {}
        except Exception as e:
            logger.error(f"Error loading product categories: {e}")
            return {}
    
    def _load_global_usage_data(self) -> Dict[str, pd.DataFrame]:
        """
        Load global usage data from CSV file.
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing the global usage data
        """
        if not self.global_usage_file:
            logger.info("No global usage file specified, skipping global usage processing")
            return {}
            
        try:
            # Read the CSV file
            df = pd.read_csv(self.global_usage_file, encoding='utf-8')
            logger.info(f"Loaded global usage data with {len(df)} rows")
            
            # Clean the data
            df = self._clean_global_usage_data(df)
            
            # Store the processed data
            self.global_usage_data['global_usage'] = df
            
            logger.info(f"Processed global usage data: {df.shape}")
            logger.info(f"Columns: {list(df.columns)}")
            
            return self.global_usage_data
            
        except Exception as e:
            logger.error(f"Error loading global usage data: {e}")
            return {}
    
    def _clean_global_usage_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and process the global usage data.
        
        Args:
            df (pd.DataFrame): Raw global usage data
            
        Returns:
            pd.DataFrame: Cleaned global usage data
        """
        if df.empty:
            return df
        
        df_clean = df.copy()
        
        # Clean ingredient names - handle the actual column name with extra spaces
        ingredient_name_col = 'Ingredient  Name'  # Note the double space
        if ingredient_name_col in df_clean.columns:
            df_clean['Ingredient Name'] = df_clean[ingredient_name_col].str.strip()
            # Remove rows with empty ingredient names
            df_clean = df_clean[df_clean['Ingredient Name'].notna() & (df_clean['Ingredient Name'] != '')]
        
        # Clean ingredient amounts - convert to numeric, handling commas
        if 'Ingredient Amount' in df_clean.columns:
            df_clean['Ingredient Amount'] = df_clean['Ingredient Amount'].astype(str)
            df_clean['Ingredient Amount'] = df_clean['Ingredient Amount'].str.replace(',', '')
            df_clean['Ingredient Amount'] = pd.to_numeric(df_clean['Ingredient Amount'], errors='coerce')
            # Fill NaN with 0
            df_clean['Ingredient Amount'] = df_clean['Ingredient Amount'].fillna(0)
        
        # Clean ingredient stock - convert to numeric, handling commas
        if 'Ingredient Stock' in df_clean.columns:
            df_clean['Ingredient Stock'] = df_clean['Ingredient Stock'].astype(str)
            df_clean['Ingredient Stock'] = df_clean['Ingredient Stock'].str.replace(',', '')
            df_clean['Ingredient Stock'] = pd.to_numeric(df_clean['Ingredient Stock'], errors='coerce')
            # Fill NaN with 0
            df_clean['Ingredient Stock'] = df_clean['Ingredient Stock'].fillna(0)
        
        logger.info(f"Cleaned global usage data: {df_clean.shape}")
        return df_clean
    
    def _create_ingredient_mapping(self) -> Dict[str, str]:
        """
        Create a mapping between ingredient names and product names from product categories.
        New ingredients will be added to product_categories.csv for manual categorization later.
        
        Returns:
            Dict[str, str]: Mapping from ingredient names to product names
        """
        if not self.global_usage_file:
            return {}
            
        try:
            # Load product categories
            product_categories = self._load_product_categories()
            
            # Create ingredient mapping
            ingredient_mapping = {}
            new_ingredients = []
            
            # Load global usage data if not already loaded
            if not self.global_usage_data:
                self._load_global_usage_data()
            
            if 'global_usage' in self.global_usage_data:
                global_df = self.global_usage_data['global_usage']
                
                for _, row in global_df.iterrows():
                    ingredient_name = row.get('Ingredient Name', '')
                    if pd.notna(ingredient_name) and ingredient_name != '':
                        # Only use exact matches
                        if ingredient_name in product_categories:
                            ingredient_mapping[ingredient_name] = ingredient_name
                            logger.debug(f"Exact match found: '{ingredient_name}' -> '{ingredient_name}'")
                        else:
                            # No exact match found, treat ingredient as a new product
                            ingredient_mapping[ingredient_name] = ingredient_name
                            new_ingredients.append(ingredient_name)
                            logger.debug(f"No exact match for ingredient '{ingredient_name}', will add as new product")
            
            self.ingredient_mapping = ingredient_mapping
            
            # Add new ingredients to product_categories.csv for manual categorization
            if new_ingredients:
                self._add_new_ingredients_to_categories(new_ingredients)
                logger.info(f"Added {len(new_ingredients)} new ingredients to product_categories.csv for manual categorization")
            
            logger.info(f"Created ingredient mapping with {len(ingredient_mapping)} entries")
            if new_ingredients:
                logger.info(f"New ingredients added: {', '.join(new_ingredients)}")
            
            return ingredient_mapping
            
        except Exception as e:
            logger.error(f"Error creating ingredient mapping: {e}")
            return {}
    
    def _add_new_ingredients_to_categories(self, new_ingredients: List[str]):
        """
        Add new ingredients to product_categories.csv with 'N/A' category for manual editing later.
        
        Args:
            new_ingredients (List[str]): List of new ingredient names to add
        """
        try:
            # Load existing categories
            existing_categories = self._load_product_categories()
            
            # Add new ingredients with 'N/A' category
            for ingredient in new_ingredients:
                if ingredient not in existing_categories:
                    existing_categories[ingredient] = 'N/A'
            
            # Save updated categories
            self._save_product_categories(existing_categories)
            
            logger.info(f"Added {len(new_ingredients)} new ingredients to product categories")
            
        except Exception as e:
            logger.error(f"Error adding new ingredients to categories: {e}")
    
    def _calculate_monthly_usage(self, ingredient_name: str) -> float:
        """
        Calculate monthly usage for a given ingredient.
        
        Args:
            ingredient_name (str): Name of the ingredient
            
        Returns:
            float: Monthly usage amount in tons (converted from kg)
        """
        try:
            if 'global_usage' not in self.global_usage_data:
                return 0.0
            
            global_df = self.global_usage_data['global_usage']
            
            # Find the ingredient in global usage data
            ingredient_row = global_df[global_df['Ingredient Name'] == ingredient_name]
            
            if not ingredient_row.empty:
                amount = ingredient_row.iloc[0].get('Ingredient Amount', 0)
                if pd.notna(amount):
                    # Convert to monthly usage (assuming the amount is annual)
                    monthly_usage_kg = amount / 12.0
                    # Convert from kg to tons by dividing by 1000
                    monthly_usage_tons = monthly_usage_kg / 1000.0
                    return monthly_usage_tons
            
            return 0.0
            
        except Exception as e:
            logger.error(f"Error calculating monthly usage for {ingredient_name}: {e}")
            return 0.0
    
    def _calculate_projected_inventory(self, total_inventory: float, monthly_usage: float, months: int) -> float:
        """
        Calculate projected inventory for a given number of months.
        
        Args:
            total_inventory (float): Total current inventory
            monthly_usage (float): Monthly usage amount
            months (int): Number of months to project
            
        Returns:
            float: Projected inventory after specified months
        """
        try:
            projected = total_inventory - (monthly_usage * months)
            return projected
        except Exception as e:
            logger.error(f"Error calculating projected inventory: {e}")
            return total_inventory
    
    def _save_product_categories(self, categories: Dict[str, str]):
        """
        Save product categories to CSV file.
        
        Args:
            categories (Dict[str, str]): Product to category mapping
        """
        try:
            import csv
            with open('config/product_categories.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Product', 'Category'])
                
                # Separate categorized and uncategorized products
                categorized = {}
                uncategorized = {}
                
                for product, category in categories.items():
                    if category and category.strip() and category.strip().upper() != 'N/A':  # Has a valid category
                        categorized[product] = category
                    else:  # No category, empty category, or N/A
                        uncategorized[product] = category
                
                # Sort categorized products alphabetically and write them first
                for product, category in sorted(categorized.items()):
                    writer.writerow([product, category])
                
                # Sort uncategorized products alphabetically and write them at the bottom
                for product, category in sorted(uncategorized.items()):
                    writer.writerow([product, category])
                
            # Log details about uncategorized/NA products
            if uncategorized:
                uncategorized_list = ', '.join(sorted(uncategorized.keys()))
                logger.info(f"Uncategorized/NA products: {uncategorized_list}")
            
            logger.info(f"Saved {len(categories)} product categories ({len(categorized)} categorized, {len(uncategorized)} uncategorized/NA)")
        except Exception as e:
            logger.error(f"Error saving product categories: {e}")
    
    def _validate_required_columns(self, df: pd.DataFrame, table_name: str) -> bool:
        """
        Validate that required product and quantity columns exist in the table.
        
        Args:
            df (pd.DataFrame): The dataframe to validate
            table_name (str): Name of the table for validation
            
        Returns:
            bool: True if all required columns exist, False otherwise
        """
        if table_name not in self.column_mapping:
            logger.warning(f"No column mapping found for table: {table_name}")
            return True  # Skip validation if no mapping exists
            
        required_columns = self.column_mapping[table_name]
        product_col = required_columns['product_column']
        quantity_col = required_columns['quantity_column']
        
        missing_columns = []
        if product_col not in df.columns:
            missing_columns.append(product_col)
        if quantity_col not in df.columns:
            missing_columns.append(quantity_col)
            
        if missing_columns:
            error_msg = f"Missing required columns for {table_name}: {missing_columns}. Available columns: {list(df.columns)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.info(f"✅ Validated required columns for {table_name}: {product_col}, {quantity_col}")
        return True
    
    def _set_table_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Programmatically find the best header row and set it as column names.
        
        Args:
            df (pd.DataFrame): The dataframe to process
            
        Returns:
            pd.DataFrame: Processed dataframe with proper headers
        """
        if df.empty:
            return df
            
        # Look for the best header row in the first 10 rows
        best_header_row = -1
        max_meaningful_headers = 0
        
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            # Count non-empty, non-numeric values that could be headers
            meaningful_count = 0
            for val in row:
                if pd.notna(val) and str(val).strip() != '':
                    val_str = str(val).strip()
                    # Check if it looks like a header (not numeric, not too long, not a data value)
                    if (not val_str.replace('.', '').replace('-', '').isdigit() and 
                        len(val_str) < 50 and
                        not val_str.replace('.', '').replace(',', '').isdigit() and
                        not any(char.isdigit() for char in val_str[:3]) and  # Avoid values that start with numbers
                        not val_str.isdigit()):  # Avoid pure numeric values
                        meaningful_count += 1
            
            if meaningful_count > max_meaningful_headers:
                max_meaningful_headers = meaningful_count
                best_header_row = i
        
        if best_header_row >= 0 and max_meaningful_headers >= 2:
            # Use this row as headers and remove all rows above it
            headers = df.iloc[best_header_row]
            # Clean header names
            clean_headers = []
            for i, header in enumerate(headers):
                if pd.notna(header) and str(header).strip() != '':
                    clean_headers.append(str(header).strip())
                else:
                    clean_headers.append(f'Column_{i}')
            
            # Create new dataframe starting from the row after the header
            new_df = df.iloc[best_header_row + 1:].copy()
            new_df.columns = clean_headers
            new_df = new_df.reset_index(drop=True)
            
            logger.info(f"Found best header row at index {best_header_row} with {max_meaningful_headers} meaningful headers, removed {best_header_row + 1} rows above")
            return new_df
        else:
            logger.warning("Could not find suitable header row, using default column names")
            return df
    
    def _handle_merged_cells(self, df: pd.DataFrame, sheet_name: str = None) -> pd.DataFrame:
        """
        Handle merged cells by propagating values.
        For UNSERVED LOCAL, only fill RAW MATERIALS cells that are part of a merged cell in Excel.
        """
        if df.empty:
            return df
        
        df_processed = df.copy()
        
        # Only apply merged cell handling to specific columns that need it
        # For UNSERVED LOCAL, only handle RAW MATERIALS column
        if sheet_name == 'UNSERVED LOCAL' and 'RAW MATERIALS' in df_processed.columns:
            col = 'RAW MATERIALS'
            # Use openpyxl to detect merged cells
            try:
                wb = openpyxl.load_workbook(self.delivery_file, data_only=True)
                ws = wb['UNSERVED LOCAL']
                merged_ranges = ws.merged_cells.ranges
                # Find the column index for RAW MATERIALS in the DataFrame
                header_row_idx = None
                for i in range(1, 11):  # search first 10 rows for header
                    if ws.cell(row=i, column=1).value and str(ws.cell(row=i, column=1).value).strip().upper() == 'RAW MATERIALS':
                        header_row_idx = i
                        break
                if header_row_idx is None:
                    logger.warning('Could not find RAW MATERIALS header row in openpyxl sheet')
                    return df_processed
                # Data starts after header
                data_start_row = header_row_idx + 1
                # Map DataFrame index to Excel row
                excel_row_for_df = lambda df_idx: data_start_row + df_idx
                # For each cell in RAW MATERIALS, only fill if part of a merged cell
                filled_values = []
                for df_idx, value in enumerate(df_processed[col]):
                    excel_row = excel_row_for_df(df_idx)
                    cell = ws.cell(row=excel_row, column=1)  # RAW MATERIALS is col 1
                    is_merged = any([cell.coordinate in rng for rng in merged_ranges])
                    if is_merged:
                        # Find the top-left cell of the merged range
                        for rng in merged_ranges:
                            if cell.coordinate in rng:
                                top_left = ws.cell(row=rng.min_row, column=rng.min_col)
                                filled_values.append(top_left.value)
                                break
                    else:
                        filled_values.append(value)
                df_processed[col] = filled_values
                # Clear RAW MATERIALS for summary rows where SUPPLIER is empty
                if 'SUPPLIER' in df_processed.columns:
                    supplier_empty_mask = df_processed['SUPPLIER'].isna() | (df_processed['SUPPLIER'] == '')
                    df_processed.loc[supplier_empty_mask, col] = ''
            except Exception as e:
                logger.warning(f'openpyxl merged cell handling failed: {e}')
        logger.info("Applied selective merged cell handling (openpyxl for UNSERVED LOCAL)")
        return df_processed
    
    def _clean_data_formatting(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean data formatting (remove .0 from integers, etc.).
        
        Args:
            df (pd.DataFrame): The dataframe to clean
            
        Returns:
            pd.DataFrame: Cleaned dataframe
        """
        if df.empty:
            return df
        
        for col in df.columns:
            if df[col].dtype == 'float64':
                # Check if all values are whole numbers
                if df[col].dropna().apply(lambda x: x.is_integer()).all():
                    df[col] = df[col].astype('Int64')
                else:
                    # Remove trailing .0 for display
                    df[col] = df[col].astype(str).str.replace('.0', '', regex=False)
        
        logger.info("Applied data formatting cleanup")
        return df
    
    def _clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean the table by removing empty rows and columns.
        
        Args:
            df (pd.DataFrame): The dataframe to clean
            
        Returns:
            pd.DataFrame: Cleaned dataframe
        """
        if df.empty:
            return df
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Fill NaN values with empty strings for string columns
        for col in df.columns:
            if hasattr(df[col], 'dtype'):
                if df[col].dtype == 'object':
                    df[col] = df[col].fillna('')
        
        return df
    
    def process_delivery_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Process a delivery sheet with standard processing.
        
        Args:
            sheet_name (str): Name of the sheet to process
            
        Returns:
            pd.DataFrame: Processed data
        """
        try:
            logger.info(f"Processing {sheet_name} sheet")
            
            # Read the sheet
            df = pd.read_excel(self.delivery_file, sheet_name=sheet_name)
            logger.info(f"Loaded {sheet_name} sheet with {len(df)} rows")
            
            # Set proper headers
            df = self._set_table_headers(df)
            
            # Handle merged cells
            df = self._handle_merged_cells(df, sheet_name)
            
            # Clean data formatting
            df = self._clean_data_formatting(df)
            
            # Clean the table
            df = self._clean_table(df)
            
            logger.info(f"Processed {sheet_name} sheet: {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            logger.error(f"Error processing {sheet_name} sheet: {e}")
            return pd.DataFrame()
    
    def process_landed_pullout_sheet(self, sheet_name: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Process a sheet that contains both LANDED and PULL data and split into two tables.
        
        Args:
            sheet_name (str): Name of the sheet to process
            
        Returns:
            Tuple[pd.DataFrame, pd.DataFrame]: Landed and Pullout tables
        """
        try:
            logger.info(f"Processing {sheet_name} sheet (detected as LANDED/PULLOUT combined)")
            
            # Read the sheet
            df = pd.read_excel(self.delivery_file, sheet_name=sheet_name)
            logger.info(f"Loaded {sheet_name} sheet with {len(df)} rows")
            
            # Find the split points
            landed_start = -1
            pullout_start = -1
            
            # Search for LANDED and PULL headers in all columns of each row
            for idx, row in df.iterrows():
                # Check all columns in the row for headers
                for col_idx, cell_val in enumerate(row):
                    if pd.notna(cell_val):
                        cell_str = str(cell_val).strip().upper()
                        
                        # Look for LANDED header (more flexible pattern)
                        if 'LANDED' in cell_str and ('HEADER' in cell_str or 'TABLE' in cell_str or col_idx == 0):
                            landed_start = idx + 1  # Start below the header
                            logger.info(f"Found 'LANDED' header at row {idx}, column {col_idx}, table starts at row {landed_start}")
                            break
                        
                        # Look for PULL header (more flexible pattern)
                        elif ('PULL' in cell_str or 'PULLOUT' in cell_str) and ('HEADER' in cell_str or 'TABLE' in cell_str or col_idx == 0):
                            pullout_start = idx + 1  # Start below the header
                            logger.info(f"Found 'PULL' header at row {idx}, column {col_idx}, table starts at row {pullout_start}")
                            break
                
                # If we found both headers, we can stop searching
                if landed_start >= 0 and pullout_start >= 0:
                    break
            
            # If no explicit LANDED header found, assume LANDED data starts from the beginning
            if landed_start == -1 and pullout_start > 0:
                landed_start = 0  # Start from the beginning
                logger.info(f"No explicit LANDED header found, assuming LANDED data starts from row 0")
            
            # Extract the tables
            if landed_start >= 0 and pullout_start > landed_start:
                # For LANDED: include all rows between headers
                landed_df = df.iloc[landed_start:pullout_start-1].copy()
                
                # For PULLOUT: include all rows after header
                pullout_df = df.iloc[pullout_start:].copy()
                
                logger.info(f"LANDED data: rows {landed_start} to {pullout_start-2}")
                logger.info(f"PULLOUT data: rows {pullout_start} to {len(df)-1}")
            else:
                logger.warning(f"Could not find proper split points for {sheet_name}")
                return pd.DataFrame(), pd.DataFrame()
            
            # Process each table
            landed_df = self._set_table_headers(landed_df)
            landed_df = self._handle_merged_cells(landed_df, 'LANDED')
            landed_df = self._clean_data_formatting(landed_df)
            landed_df = self._clean_table(landed_df)
            
            pullout_df = self._set_table_headers(pullout_df)
            pullout_df = self._handle_merged_cells(pullout_df, 'PULLOUT')
            pullout_df = self._clean_data_formatting(pullout_df)
            pullout_df = self._clean_table(pullout_df)
            
            # Validate columns
            self._validate_required_columns(landed_df, 'LANDED')
            self._validate_required_columns(pullout_df, 'PULLOUT')
            
            logger.info(f"Processed LANDED table: {len(landed_df)} rows")
            logger.info(f"Processed PULLOUT table: {len(pullout_df)} rows")
            
            return landed_df, pullout_df
            
        except Exception as e:
            logger.error(f"Error processing {sheet_name} sheet: {e}")
            return pd.DataFrame(), pd.DataFrame()
    
    def process_inventory_sheet(self) -> pd.DataFrame:
        """
        Process the inventory sheet.
        
        Returns:
            pd.DataFrame: Processed inventory data
        """
        try:
            logger.info("Processing inventory sheet")
            
            # Read the inventory file
            df = pd.read_excel(self.inventory_file)
            logger.info(f"Loaded inventory sheet with {len(df)} rows")
            
            # Find the header row (look for "ITEM DESCRIPTION" and "QTY IN KGS")
            header_row = -1
            for i in range(min(25, len(df))):
                row_values = [str(val).strip() if pd.notna(val) else '' for val in df.iloc[i]]
                if 'ITEM DESCRIPTION' in row_values and 'QTY IN KGS' in row_values:
                    header_row = i
                    break
            
            if header_row == -1:
                logger.warning("Could not find inventory header row")
                return pd.DataFrame()
            
            logger.info(f"Found header row at index {header_row}")
            
            # Extract headers
            headers = df.iloc[header_row]
            clean_headers = []
            for i, header in enumerate(headers):
                if pd.notna(header) and str(header).strip() != '':
                    clean_headers.append(str(header).strip())
                else:
                    clean_headers.append(f'Column_{i}')
            
            # Start data from the row after header
            data_df = df.iloc[header_row + 1:].copy()
            data_df.columns = clean_headers
            
            # Clean the data
            data_df = data_df.dropna(how='all')  # Remove completely empty rows
            
            # Extract only the relevant columns (ITEM DESCRIPTION and QTY IN KGS)
            relevant_columns = []
            for col in data_df.columns:
                if 'ITEM DESCRIPTION' in col or 'QTY IN KGS' in col or 'DESCRIPTION' in col or 'QTY' in col:
                    relevant_columns.append(col)
            
            if relevant_columns:
                data_df = data_df[relevant_columns]
            
            # Clean up the data
            for col in data_df.columns:
                if data_df[col].dtype == 'object':
                    data_df[col] = data_df[col].fillna('')
            
            # Remove rows that are completely empty or contain only formatting
            data_df = data_df[data_df.apply(lambda row: row.astype(str).str.strip().ne('').any(), axis=1)]
            
            logger.info(f"Processed inventory sheet: {len(data_df)} rows, {len(data_df.columns)} columns")
            return data_df
            
        except Exception as e:
            logger.error(f"Error processing inventory sheet: {e}")
            return pd.DataFrame()
    
    def process_all_data(self) -> Dict[str, pd.DataFrame]:
        """
        Process all data including delivery, inventory, and global usage if available.
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing all processed data
        """
        # Process delivery data
        self._process_delivery_data()
        
        # Process inventory data
        self._process_inventory_data()
        
        # Process global usage data if available
        if self.global_usage_file:
            self._load_global_usage_data()
        
        # Combine all results
        all_results = {**self.delivery_data, **self.inventory_data, **self.global_usage_data}
        
        logger.info(f"Processed all data: {len(all_results)} datasets")
        return all_results
    
    def _process_delivery_data(self):
        """
        Process all delivery data sheets.
        """
        # Load delivery file
        delivery_excel = pd.ExcelFile(self.delivery_file)
        logger.info(f"Available delivery sheets: {delivery_excel.sheet_names}")
        
        # Process delivery sheets
        for sheet_name in delivery_excel.sheet_names:
            if sheet_name == 'SUMMARY':
                continue  # Skip SUMMARY sheet
            elif 'LANDED' in sheet_name.upper() and 'PULL' in sheet_name.upper():
                # Auto-detect and split LANDED/PULLOUT combined sheets
                landed_df, pullout_df = self.process_landed_pullout_sheet(sheet_name)
                if not landed_df.empty:
                    self.delivery_data['LANDED'] = landed_df
                if not pullout_df.empty:
                    self.delivery_data['PULLOUT'] = pullout_df
            else:
                # Standard processing for other sheets
                df = self.process_delivery_sheet(sheet_name)
                if not df.empty:
                    # Validate columns if mapping exists
                    if sheet_name in self.column_mapping:
                        self._validate_required_columns(df, sheet_name)
                    self.delivery_data[sheet_name] = df
    
    def _process_inventory_data(self):
        """
        Process inventory data.
        """
        inventory_df = self.process_inventory_sheet()
        if not inventory_df.empty:
            self.inventory_data['INVENTORY'] = inventory_df
    
    def combine_all_data(self) -> pd.DataFrame:
        """
        Combine all tables (except SUMMARY) including inventory data and global usage data.
        
        Returns:
            pd.DataFrame: Combined dataframe with all data
        """
        combined_data = []
        
        # Load product categories
        categories = self._load_product_categories()
        new_products = set()
        blank_categories = []
        
        # Process delivery data
        for table_name, df in self.delivery_data.items():
            if df.empty:
                continue
                
            # Get the required columns for this table
            # Handle table name variations (spaces vs underscores)
            mapping_key = table_name
            if table_name not in self.column_mapping:
                # Try with underscores
                mapping_key = table_name.replace(' ', '_')
                if mapping_key not in self.column_mapping:
                    logger.warning(f"No column mapping found for table: {table_name} or {mapping_key}")
                    continue
            
            required_columns = self.column_mapping[mapping_key]
            product_col = required_columns['product_column']
            quantity_col = required_columns['quantity_column']
            
            # Check if required columns exist
            if product_col not in df.columns or quantity_col not in df.columns:
                logger.warning(f"Missing required columns for {table_name}: {product_col}, {quantity_col}")
                continue
            
            # Extract product and quantity data
            table_data = df[[product_col, quantity_col]].copy()
            
            # Add status column
            table_data['STATUS'] = table_name.replace('_', ' ')
            
            # Convert UNSERVED LOCAL quantity from kilos to tons (divide by 1000)
            # Use fuzzy matching to handle different variations (UNSERVED LOCAL, UNSERVED_LOCAL, etc.)
            if 'UNSERVED' in table_name.upper() and 'LOCAL' in table_name.upper():
                # Convert to numeric, handle errors, then divide by 1000 to convert kilos to tons
                table_data[quantity_col] = pd.to_numeric(table_data[quantity_col], errors='coerce')
                table_data[quantity_col] = table_data[quantity_col] / 1000
                logger.info(f"Converted {table_name} quantities from kilos to tons (divided by 1000)")
            
            # Rename columns to standard names
            table_data = table_data.rename(columns={
                product_col: 'PRODUCT',
                quantity_col: 'QUANTITY'
            })
            
            # Remove rows where product is empty
            table_data = table_data.dropna(subset=['PRODUCT'])
            table_data = table_data[table_data['PRODUCT'] != '']
            
            # Filter out summary rows only
            # Remove rows where product is empty or contains only whitespace
            table_data = table_data[table_data['PRODUCT'].str.strip() != '']
            
            logger.info(f"Filtered {table_name} data: removed empty product rows")
            
            # Add category column and track new/blank categories
            table_data['CATEGORY'] = ''
            for idx, row in table_data.iterrows():
                product = row['PRODUCT'].strip()
                
                # Check if product exists in categories
                if product in categories:
                    category = categories[product]
                    table_data.at[idx, 'CATEGORY'] = category
                    
                    # Track blank categories
                    if not category:
                        blank_categories.append(product)
                        logger.debug(f"Product '{product}' has blank category")
                    elif category.strip().upper() == 'N/A':
                        logger.debug(f"Product '{product}' has N/A category")
                    else:
                        logger.debug(f"Product '{product}' categorized as: {category}")
                else:
                    # New product found
                    new_products.add(product)
                    table_data.at[idx, 'CATEGORY'] = ''
                    logger.debug(f"New product found: '{product}' (no category assigned)")
            
            combined_data.append(table_data)
            logger.info(f"Added {len(table_data)} rows from {table_name}")
        
        # Process inventory data
        if 'INVENTORY' in self.inventory_data:
            inventory_df = self.inventory_data['INVENTORY']
            if not inventory_df.empty:
                # Rename inventory columns to match format
                inventory_data = inventory_df.copy()
                inventory_data = inventory_data.rename(columns={
                    'ITEM DESCRIPTION': 'PRODUCT',
                    'QTY IN KGS': 'QUANTITY'
                })
                
                # Divide inventory quantity by 1000
                inventory_data['QUANTITY'] = pd.to_numeric(inventory_data['QUANTITY'], errors='coerce')
                inventory_data['QUANTITY'] = inventory_data['QUANTITY'] / 1000
                logger.info(f"Divided INVENTORY quantities by 1000")
                
                # Add status column
                inventory_data['STATUS'] = 'INVENTORY'
                
                # Remove rows where product is empty
                inventory_data = inventory_data.dropna(subset=['PRODUCT'])
                inventory_data = inventory_data[inventory_data['PRODUCT'] != '']
                
                # Add category column and track new/blank categories
                inventory_data['CATEGORY'] = ''
                for idx, row in inventory_data.iterrows():
                    product = row['PRODUCT'].strip()
                    
                    # Check if product exists in categories
                    if product in categories:
                        category = categories[product]
                        inventory_data.at[idx, 'CATEGORY'] = category
                        
                        # Track blank categories
                        if not category:
                            blank_categories.append(product)
                    else:
                        # New product found
                        new_products.add(product)
                        inventory_data.at[idx, 'CATEGORY'] = ''
                
                combined_data.append(inventory_data)
                logger.info(f"Added {len(inventory_data)} rows from INVENTORY")
        
        # Process global usage data if available
        if self.global_usage_file and 'global_usage' in self.global_usage_data:
            global_df = self.global_usage_data['global_usage']
            if not global_df.empty:
                # Create global usage data with negative quantities (representing usage)
                global_usage_data = []
                
                for _, row in global_df.iterrows():
                    ingredient_name = row.get('Ingredient Name', '')
                    ingredient_amount = row.get('Ingredient Amount', 0)
                    
                    if pd.notna(ingredient_name) and ingredient_name != '' and pd.notna(ingredient_amount):
                        # Convert monthly usage from kg to tons (values are already monthly)
                        monthly_usage_tons = ingredient_amount / 1000.0  # Convert kg to tons
                        
                        # Find category for this ingredient
                        category = ''
                        if ingredient_name in categories:
                            category = categories[ingredient_name]
                        else:
                            # If no category found, use ingredient name as category
                            category = ingredient_name
                            # Add to product categories for manual editing later
                            if ingredient_name not in categories:
                                new_products.add(ingredient_name)
                                categories[ingredient_name] = ''
                        
                        global_usage_data.append({
                            'PRODUCT': ingredient_name,
                            'QUANTITY': monthly_usage_tons,
                            'STATUS': 'GLOBAL USAGE',
                            'CATEGORY': category
                        })
                
                if global_usage_data:
                    global_df_combined = pd.DataFrame(global_usage_data)
                    combined_data.append(global_df_combined)
                    logger.info(f"Added {len(global_usage_data)} rows from GLOBAL USAGE")
        
        if combined_data:
            # Combine all dataframes
            final_df = pd.concat(combined_data, ignore_index=True)
            logger.info(f"Combined data: {len(final_df)} total rows from {len(combined_data)} tables")
            
            # Handle new products and warnings
            if new_products:
                logger.info(f"Found {len(new_products)} new products, adding to categories file")
                for product in new_products:
                    categories[product] = ''
                self._save_product_categories(categories)
                
                logger.warning(f"NEW PRODUCTS ADDED TO product_categories.csv: {sorted(new_products)}")
                print(f"\n⚠️  NEW PRODUCTS ADDED TO product_categories.csv:")
                for product in sorted(new_products):
                    print(f"  - {product}")
                print("Please edit product_categories.csv to add categories for these products.")
            
            if blank_categories:
                unique_blank = list(set(blank_categories))
                logger.warning(f"Found products with blank categories: {sorted(unique_blank)}")
                print(f"\n⚠️  WARNING: Found products with blank categories:")
                for product in sorted(unique_blank):
                    print(f"  - {product}")
                print("Please edit product_categories.csv to add categories for these products.")
            
            # Log category summary
            if not final_df.empty:
                category_counts = final_df['CATEGORY'].value_counts()
                logger.info("Category breakdown in combined data:")
                for category, count in category_counts.items():
                    if category and category.strip():
                        logger.info(f"  {category}: {count} products")
                    else:
                        logger.info(f"  No category: {count} products")
                
                # Check for N/A categories specifically
                na_count = len(final_df[final_df['CATEGORY'] == 'N/A'])
                if na_count > 0:
                    na_products = final_df[final_df['CATEGORY'] == 'N/A']['PRODUCT'].unique()
                    logger.info(f"N/A category products ({na_count}): {', '.join(sorted(na_products))}")
                
                # Log status breakdown including global usage
                status_counts = final_df['STATUS'].value_counts()
                logger.info("Status breakdown in combined data:")
                for status, count in status_counts.items():
                    logger.info(f"  {status}: {count} products")
            
            return final_df
        else:
            logger.warning("No data to combine")
            return pd.DataFrame(columns=['PRODUCT', 'QUANTITY', 'STATUS', 'CATEGORY'])
    
    def save_processed_data(self, output_dir: str = None):
        """
        Save all processed data to CSV files.
        
        Args:
            output_dir (str): Directory to save the processed data. If None, uses self.output_dir
        """
        import os
        
        # Use instance output directory if none specified
        if output_dir is None:
            output_dir = self.output_dir
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Saving processed data to: {output_dir}")
        
        # Save delivery data
        for sheet_name, df in self.delivery_data.items():
            if not df.empty:
                output_file = os.path.join(output_dir, f"{sheet_name}.csv")
                df.to_csv(output_file, index=False)
                logger.info(f"Saved {sheet_name} to {output_file}")
        
        # Save inventory data
        for sheet_name, df in self.inventory_data.items():
            if not df.empty:
                output_file = os.path.join(output_dir, "inventory.csv")
                df.to_csv(output_file, index=False)
                logger.info(f"Saved {sheet_name} to {output_file}")
    
    def get_summary(self) -> Dict[str, dict]:
        """
        Get a summary of all processed data.
        
        Returns:
            Dict[str, dict]: Summary information for each table
        """
        summary = {}
        
        # Delivery data summary
        for sheet_name, df in self.delivery_data.items():
            if not df.empty:
                summary[sheet_name] = {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': list(df.columns),
                    'data_types': df.dtypes.to_dict(),
                    'missing_values': df.isnull().sum().to_dict()
                }
            else:
                summary[sheet_name] = {
                    'rows': 0,
                    'columns': 0,
                    'column_names': [],
                    'data_types': {},
                    'missing_values': {}
                }
        
        # Inventory data summary
        for sheet_name, df in self.inventory_data.items():
            if not df.empty:
                summary[sheet_name] = {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': list(df.columns),
                    'data_types': df.dtypes.to_dict(),
                    'missing_values': df.isnull().sum().to_dict()
                }
            else:
                summary[sheet_name] = {
                    'rows': 0,
                    'columns': 0,
                    'column_names': [],
                    'data_types': {},
                    'missing_values': {}
                }
        
        return summary

    def create_pivot_table(self, combined_df: pd.DataFrame = None) -> pd.DataFrame:
        """
        Create a pivot table with Category in rows, Status in columns, and sum of quantities as values.
        Global Usage is positioned to the right of the TOTAL column and is not included in total calculations.
        
        Args:
            combined_df (pd.DataFrame, optional): Combined dataframe. If None, will use self.combine_all_data()
            
        Returns:
            pd.DataFrame: Pivot table
        """
        if combined_df is None:
            combined_df = self.combine_all_data()
        
        if combined_df.empty:
            logger.warning("No data available for pivot table")
            return pd.DataFrame()
        
        # Ensure QUANTITY is numeric
        combined_df['QUANTITY'] = pd.to_numeric(combined_df['QUANTITY'], errors='coerce')
        
        # Fill NaN values in CATEGORY with 'Uncategorized'
        combined_df['CATEGORY'] = combined_df['CATEGORY'].fillna('Uncategorized')
        combined_df['CATEGORY'] = combined_df['CATEGORY'].replace('', 'Uncategorized')
        
        # Filter out products with 'N/A' category for pivot table
        original_count = len(combined_df)
        combined_df_filtered = combined_df[combined_df['CATEGORY'] != 'N/A'].copy()
        filtered_count = len(combined_df_filtered)
        excluded_count = original_count - filtered_count
        
        if excluded_count > 0:
            # Get the specific products with N/A category
            na_products = combined_df[combined_df['CATEGORY'] == 'N/A']['PRODUCT'].unique()
            na_products_list = ', '.join(sorted(na_products))
            
            logger.info(f"Excluded {excluded_count} products with 'N/A' category from pivot table")
            logger.info(f"N/A products: {na_products_list}")
            logger.info(f"Pivot table will show {filtered_count} products (out of {original_count} total)")
        else:
            logger.info(f"All {original_count} products have valid categories for pivot table")
        
        # Create pivot table using filtered data
        try:
            pivot_table = pd.pivot_table(
                combined_df_filtered,
                values='QUANTITY',
                index='CATEGORY',
                columns='STATUS',
                aggfunc='sum',
                fill_value=0
            )
            
            # Add empty TOTAL column (will be filled with formulas in Excel)
            pivot_table['TOTAL'] = 0
            
            # Add empty TOTAL row (will be filled with formulas in Excel)
            pivot_table.loc['TOTAL'] = 0
            
            logger.info(f"Created pivot table with shape: {pivot_table.shape}")
            logger.info(f"Categories: {list(pivot_table.index[:-1])}")  # Exclude 'TOTAL' row
            logger.info(f"Statuses: {list(pivot_table.columns[:-1])}")  # Exclude 'TOTAL' column
            
            return pivot_table
            
        except Exception as e:
            logger.error(f"Error creating pivot table: {e}")
            return pd.DataFrame()
    
    def create_enhanced_pivot_table(self, combined_df: pd.DataFrame = None) -> pd.DataFrame:
        """
        Create an enhanced pivot table with monthly usage and projected inventory.
        Global Usage is positioned to the right of the TOTAL column and is not included in total calculations.
        
        Args:
            combined_df (pd.DataFrame, optional): Combined dataframe. If None, will use self.combine_all_data()
            
        Returns:
            pd.DataFrame: Enhanced pivot table with usage and projections
        """
        if combined_df is None:
            combined_df = self.combine_all_data()
        
        if combined_df.empty:
            logger.warning("No data available for enhanced pivot table")
            return pd.DataFrame()
        
        # Create the base pivot table
        base_pivot = self.create_pivot_table(combined_df)
        
        if base_pivot.empty:
            logger.warning("No base pivot table available for enhancement")
            return pd.DataFrame()
        
        # Create enhanced pivot table
        enhanced_pivot = base_pivot.copy()
        
        # Add Monthly Usage column
        enhanced_pivot['Monthly Usage'] = 0.0
        
        # Add projection columns
        for month in range(1, 7):
            enhanced_pivot[f'{month} Month Projection'] = 0.0
        
        # Process global usage data if available
        if self.global_usage_file and 'global_usage' in self.global_usage_data:
            # Get monthly usage from the combined data (GLOBAL USAGE status)
            global_usage_df = combined_df[combined_df['STATUS'] == 'GLOBAL USAGE']
            
            if not global_usage_df.empty:
                # Create category-based usage mapping
                category_usage = {}
                
                for _, row in global_usage_df.iterrows():
                    ingredient_name = row['PRODUCT']
                    monthly_usage = row['QUANTITY']
                    category = row['CATEGORY']
                    
                    # Sum usage by category
                    if category in category_usage:
                        category_usage[category] += monthly_usage
                    else:
                        category_usage[category] = monthly_usage
                
                # Add monthly usage to existing categories or create new category rows
                for category, total_usage in category_usage.items():
                    if category in enhanced_pivot.index:
                        # Category exists, add monthly usage
                        enhanced_pivot.loc[category, 'Monthly Usage'] = total_usage
                    else:
                        # Category doesn't exist, create new row
                        new_row = pd.Series(0.0, index=enhanced_pivot.columns)
                        new_row['Monthly Usage'] = total_usage
                        enhanced_pivot.loc[category] = new_row
        
        # Calculate projections for all categories
        for category in enhanced_pivot.index:
            if category != 'TOTAL':
                monthly_usage = enhanced_pivot.loc[category, 'Monthly Usage']
                total_inventory = enhanced_pivot.loc[category, 'TOTAL']
                
                # Calculate projections by subtracting monthly usage
                for month in range(1, 7):
                    if total_inventory == "-" or pd.isna(monthly_usage):
                        # If no monthly usage, projection equals total inventory
                        projection = - (monthly_usage * month)
                    else:
                        # Calculate projection by subtracting monthly usage
                        projection = total_inventory - (monthly_usage * month)
                    enhanced_pivot.loc[category, f'{month} Month Projection'] = projection
        
        # Ensure the TOTAL row has proper values for all columns
        # Calculate row totals (sum across all status columns for each category, excluding GLOBAL USAGE)
        for category in enhanced_pivot.index:
            if category != 'TOTAL':
                # Sum all status columns (excluding TOTAL, GLOBAL USAGE, Monthly Usage, and projection columns)
                status_columns = [col for col in enhanced_pivot.columns 
                                if col not in ['TOTAL', 'GLOBAL USAGE', 'Monthly Usage'] 
                                and 'Month Projection' not in col]
                row_total = enhanced_pivot.loc[category, status_columns].sum()
                enhanced_pivot.loc[category, 'TOTAL'] = row_total
        
        # Calculate column totals (sum down all rows for each status)
        for col in enhanced_pivot.columns:
            if col != 'TOTAL':
                # Sum all rows (excluding TOTAL row)
                col_total = enhanced_pivot.loc[enhanced_pivot.index != 'TOTAL', col].sum()
                enhanced_pivot.loc['TOTAL', col] = col_total
        
        logger.info(f"Created enhanced pivot table with {len(enhanced_pivot)} categories")
        logger.info(f"Column order: {list(enhanced_pivot.columns)}")
        return enhanced_pivot
    
    def save_pivot_table(self, pivot_df: pd.DataFrame, output_dir: str = None):
        """
        Save the pivot table to CSV and Excel files.
        
        Args:
            pivot_df (pd.DataFrame): Pivot table to save
            output_dir (str): Directory to save the pivot table. If None, uses self.output_dir
        """
        if pivot_df.empty:
            logger.warning("No pivot table to save")
            return
        
        # Use instance output directory if none specified
        if output_dir is None:
            output_dir = self.output_dir
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Saving pivot table to: {output_dir}")
        
        # Save to CSV
        csv_file = os.path.join(output_dir, "PIVOT_TABLE.csv")
        pivot_df.to_csv(csv_file)
        logger.info(f"Saved pivot table to {csv_file}")
        
        # Save to Excel with formatting
        excel_file = os.path.join(output_dir, "PIVOT_TABLE.xlsx")
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='Pivot Table')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Pivot Table']
            
            # Format the worksheet
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Format row headers (categories)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.border = border
                
                # Highlight total row
                if cell.value == 'TOTAL':
                    for col in range(1, worksheet.max_column + 1):
                        total_cell = worksheet.cell(row=row, column=col)
                        total_cell.font = total_font
                        total_cell.fill = total_fill
                        total_cell.border = border
            
            # Format all data cells and replace zeros with dashes
            for row in range(2, worksheet.max_row + 1):
                for col in range(2, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
                    
                    # Get column header to determine formatting
                    col_header = worksheet.cell(row=1, column=col).value
                    
                    # Format GLOBAL USAGE column
                    if col_header == 'GLOBAL USAGE':
                        # Keep actual values as 0 but display as dash
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        else:
                            # Format negative values (usage) with proper number format
                            cell.number_format = '#,##0.00'
                            # Apply red color for negative values
                            if cell.value < 0:
                                cell.font = Font(color="FF0000", bold=True)
                    
                    # Format other columns
                    else:
                        # Replace zero values with dash for display
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")  # Gray color for dashes
                        elif isinstance(cell.value, (int, float)) and cell.value > 0:
                            # Format all numbers with comma separation and 0 decimal places
                            cell.number_format = '#,##0'
            
            # Add SUM formulas for totals
            # Row totals (sum across columns for each category)
            for row in range(2, worksheet.max_row):  # Exclude the TOTAL row
                # Calculate the range for this row (from column B to the column before TOTAL)
                start_col = get_column_letter(2)  # Column B
                # Find the TOTAL column (it should be the second-to-last column)
                total_col = worksheet.max_column - 1
                end_col = get_column_letter(total_col - 1)  # Column before TOTAL
                row_num = row
                
                # Create SUM formula for row total (excluding GLOBAL USAGE)
                sum_formula = f"=SUM({start_col}{row_num}:{end_col}{row_num})"
                total_cell = worksheet.cell(row=row, column=total_col)
                total_cell.value = sum_formula
                total_cell.number_format = '#,##0'
                total_cell.font = Font(bold=True)
                total_cell.border = border
                total_cell.alignment = Alignment(horizontal='right')
            
            # Column totals (sum down rows for each status)
            for col in range(2, total_col):  # Exclude the TOTAL column and GLOBAL USAGE column
                # Calculate the range for this column (from row 2 to the second-to-last row)
                col_letter = get_column_letter(col)
                start_row = 2
                end_row = worksheet.max_row - 1  # Second-to-last row
                
                # Create SUM formula for column total
                sum_formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                total_cell = worksheet.cell(row=worksheet.max_row, column=col)
                total_cell.value = sum_formula
                total_cell.number_format = '#,##0'
                total_cell.font = total_font
                total_cell.fill = total_fill
                total_cell.border = border
                total_cell.alignment = Alignment(horizontal='right')
            
            # Grand total (sum of all data cells excluding GLOBAL USAGE)
            start_col = get_column_letter(2)  # Column B
            end_col = get_column_letter(total_col - 1)  # Column before TOTAL
            start_row = 2
            end_row = worksheet.max_row - 1  # Second-to-last row
            
            grand_total_formula = f"=SUM({start_col}{start_row}:{end_col}{end_row})"
            grand_total_cell = worksheet.cell(row=worksheet.max_row, column=total_col)
            grand_total_cell.value = grand_total_formula
            grand_total_cell.number_format = '#,##0'
            grand_total_cell.font = total_font
            grand_total_cell.fill = total_fill
            grand_total_cell.border = border
            grand_total_cell.alignment = Alignment(horizontal='right')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Saved formatted pivot table to {excel_file}")
    
    def save_enhanced_pivot_table(self, enhanced_pivot_df: pd.DataFrame, output_dir: str = None):
        """
        Save the enhanced pivot table to CSV and Excel files with formatting.
        
        Args:
            enhanced_pivot_df (pd.DataFrame): Enhanced pivot table to save
            output_dir (str): Directory to save the pivot table. If None, uses self.output_dir
        """
        if enhanced_pivot_df.empty:
            logger.warning("No enhanced pivot table to save")
            return
        
        # Use instance output directory if none specified
        if output_dir is None:
            output_dir = self.output_dir
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Saving enhanced pivot table to: {output_dir}")
        
        # Save to CSV
        csv_file = os.path.join(output_dir, "ENHANCED_PIVOT_TABLE.csv")
        enhanced_pivot_df.to_csv(csv_file)
        logger.info(f"Saved enhanced pivot table to {csv_file}")
        
        # Save to Excel with enhanced formatting
        excel_file = os.path.join(output_dir, "ENHANCED_PIVOT_TABLE.xlsx")
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            enhanced_pivot_df.to_excel(writer, sheet_name='Enhanced Pivot Table')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Enhanced Pivot Table']
            
            # Format the worksheet
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            usage_font = Font(bold=True, color="000000")
            usage_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            projection_font = Font(bold=True, color="000000")
            projection_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            negative_font = Font(bold=True, color="FF0000")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Format row headers (categories)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.border = border
                
                # Highlight total row
                if cell.value == 'TOTAL':
                    for col in range(1, worksheet.max_column + 1):
                        total_cell = worksheet.cell(row=row, column=col)
                        total_cell.font = total_font
                        total_cell.fill = total_fill
                        total_cell.border = border
            
            # Format all data cells
            for row in range(2, worksheet.max_row + 1):
                for col in range(2, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
                    
                    # Get column header to determine formatting
                    col_header = worksheet.cell(row=1, column=col).value
                    
                    # Format Monthly Usage column
                    if col_header == 'Monthly Usage':
                        cell.fill = usage_fill
                        cell.font = usage_font
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        else:
                            cell.number_format = '#,##0.00'
                    
                    # Format projection columns
                    elif 'Month Projection' in str(col_header):
                        cell.fill = projection_fill
                        cell.font = projection_font
                        
                        # Add Excel formula for projections
                        if cell.value != 0 and cell.value != 0.0:
                            # Get the row number for this category
                            category_row = row
                            
                            # Find the TOTAL column position
                            total_col_idx = None
                            for col_idx in range(1, worksheet.max_column + 1):
                                if worksheet.cell(row=1, column=col_idx).value == 'TOTAL':
                                    total_col_idx = col_idx
                                    break
                            
                            # Find the Monthly Usage column position
                            usage_col_idx = None
                            for col_idx in range(1, worksheet.max_column + 1):
                                if worksheet.cell(row=1, column=col_idx).value == 'Monthly Usage':
                                    usage_col_idx = col_idx
                                    break
                            
                            if total_col_idx and usage_col_idx:
                                total_col_letter = get_column_letter(total_col_idx)
                                usage_col_letter = get_column_letter(usage_col_idx)
                                
                                # Extract month number from column header
                                month_num = int(str(col_header).split()[0])
                                
                                # Create formula with IF statement to handle special cases:
                                # =IF(total_col="-",-(usage_col*month),IF(usage_col="-",total_col,total_col-(usage_col*month)))
                                formula = f'=IF({total_col_letter}{category_row}="-",-({usage_col_letter}{category_row}*{month_num}),IF({usage_col_letter}{category_row}="-",{total_col_letter}{category_row},{total_col_letter}{category_row}-({usage_col_letter}{category_row}*{month_num})))'
                                cell.value = formula
                                
                                # Check if the calculated value would be negative and apply red formatting
                                # We'll need to calculate this for formatting purposes
                                try:
                                    total_value = enhanced_pivot_df.iloc[category_row-2, enhanced_pivot_df.columns.get_loc('TOTAL')]
                                    usage_value = enhanced_pivot_df.iloc[category_row-2, enhanced_pivot_df.columns.get_loc('Monthly Usage')]
                                    if usage_value == 0 or pd.isna(usage_value):
                                        calculated_value = total_value
                                    else:
                                        calculated_value = total_value - (usage_value * month_num)
                                    
                                    if calculated_value < 0:
                                        cell.font = negative_font
                                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                                except:
                                    pass
                        
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        else:
                            cell.number_format = '#,##0.00'
                    
                    # Format GLOBAL USAGE column
                    elif col_header == 'GLOBAL USAGE':
                        # Keep actual values as 0 but display as dash
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        else:
                            # Format negative values (usage) with proper number format
                            cell.number_format = '#,##0.00'
                            # Apply red color for negative values
                            if cell.value < 0:
                                cell.font = Font(color="FF0000", bold=True)
                    
                    # Format regular data cells
                    else:
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        elif isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell.number_format = '#,##0'
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Saved enhanced formatted pivot table to {excel_file}")


def main():
    """
    Main function to demonstrate the Combined processor usage.
    """
    import os
    
    # Check if global usage file exists
    global_usage_file = None
    if os.path.exists('input/global_usage.csv'):
        global_usage_file = 'input/global_usage.csv'
        print("🌍 Global usage file found - will create enhanced pivot table with projections")
    else:
        print("📊 Running in standard mode (no global usage file)")
    
    # Initialize the processor
    processor = CombinedProcessor('input/deliveries.xlsx', 'input/inventory.xlsx', global_usage_file)
    
    # Process all data
    results = processor.process_all_data()
    
    # Ensure global usage data is loaded if available
    if global_usage_file:
        processor._load_global_usage_data()
        print(f"🌍 Global usage data loaded: {len(processor.global_usage_data.get('global_usage', pd.DataFrame()))} ingredients")
    
    # Print summary
    summary = processor.get_summary()
    print("\n=== COMBINED PROCESSING SUMMARY ===")
    for sheet_name, info in summary.items():
        print(f"\n{sheet_name.upper()}:")
        print(f"  Rows: {info['rows']}")
        print(f"  Columns: {info['columns']}")
        if info['column_names']:
            print(f"  Columns: {', '.join(info['column_names'])}")
    
    # Save processed data
    processor.save_processed_data()
    
    # Combine all data
    combined_df = processor.combine_all_data()
    
    # Save combined data
    if not combined_df.empty:
        combined_file = os.path.join(processor.output_dir, "COMBINED_ALL_DATA.csv")
        combined_df.to_csv(combined_file, index=False)
        logger.info(f"Saved combined data to {combined_file}")
        print(f"\n=== COMBINED DATA SUMMARY ===")
        print(f"Total rows: {len(combined_df)}")
        print(f"Status breakdown: {combined_df['STATUS'].value_counts().to_dict()}")
        print(f"Category breakdown: {combined_df['CATEGORY'].value_counts().to_dict()}")
        print(f"Sample data:")
        print(combined_df.head(10).to_string(index=False))
        
        # Create and save pivot table
        print(f"\n=== CREATING PIVOT TABLE ===")
        
        # Check for N/A categories before creating pivot table
        na_count = len(combined_df[combined_df['CATEGORY'] == 'N/A'])
        if na_count > 0:
            na_products = combined_df[combined_df['CATEGORY'] == 'N/A']['PRODUCT'].unique()
            print(f"⚠️  Found {na_count} products with 'N/A' category - these will be excluded from pivot table")
            print(f"📋 N/A products: {', '.join(sorted(na_products))}")
        
        pivot_table = processor.create_pivot_table(combined_df)
        
        if not pivot_table.empty:
            processor.save_pivot_table(pivot_table)
            print(f"✅ Standard pivot table created successfully!")
            print(f"Shape: {pivot_table.shape}")
            print(f"Categories: {len(pivot_table.index) - 1}")  # Exclude TOTAL row
            # Count statuses excluding TOTAL and GLOBAL USAGE columns
            status_columns = [col for col in pivot_table.columns if col not in ['TOTAL', 'GLOBAL USAGE']]
            print(f"Statuses: {len(status_columns)}")
            print(f"Column order: {list(pivot_table.columns)}")
            print(f"\nStandard Pivot Table Preview:")
            print(pivot_table.to_string())
        else:
            print("❌ Failed to create standard pivot table")
        
        # Create enhanced pivot table if global usage file is available
        if global_usage_file:
            print(f"\n=== CREATING ENHANCED PIVOT TABLE ===")
            
            enhanced_pivot_table = processor.create_enhanced_pivot_table(combined_df)
            
            if not enhanced_pivot_table.empty:
                processor.save_enhanced_pivot_table(enhanced_pivot_table)
                print(f"✅ Enhanced pivot table created successfully!")
                print(f"Shape: {enhanced_pivot_table.shape}")
                print(f"Categories: {len(enhanced_pivot_table.index) - 1}")  # Exclude TOTAL row
                print(f"Columns: {len(enhanced_pivot_table.columns)}")
                print(f"Column order: {list(enhanced_pivot_table.columns)}")
                print(f"\nEnhanced Pivot Table Summary:")
                print(f"   GLOBAL USAGE column (right of TOTAL): ✅")
                print(f"   Monthly Usage column: ✅")
                print(f"   1-6 Month Projections: ✅")
                print(f"   Negative values colored red: ✅")
                print(f"\nEnhanced Pivot Table Preview:")
                print(enhanced_pivot_table.to_string())
            else:
                print("❌ Failed to create enhanced pivot table")
    
    print(f"\n=== INDIVIDUAL DATA SAMPLES ===")
    for sheet_name, df in results.items():
        if not df.empty:
            print(f"\n{sheet_name}:")
            print(f"Shape: {df.shape}")
            print("First 5 rows:")
            print(df.head().to_string(index=False))
            print("\n" + "="*50)


if __name__ == "__main__":
    main() 