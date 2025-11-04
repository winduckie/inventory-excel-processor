#!/usr/bin/env python3
"""
Global Usage Processor

A processor that integrates global usage data with the existing combined processor system.
Adds monthly usage and projected inventory calculations to the pivot table.
"""

import pandas as pd
import numpy as np
import json
from typing import Dict, List, Tuple, Optional
import logging
import os
from datetime import datetime
import openpyxl
from combined_processor import CombinedProcessor

# Set up logging
os.makedirs('logs', exist_ok=True)

# Create a unique log filename with timestamp
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
log_filename = f'logs/global_usage_processor_{timestamp}.log'

# Set up logging to both file and console
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, mode='w'),
        logging.StreamHandler()  # This keeps console output
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"Logging to file: {log_filename}")


class GlobalUsageProcessor(CombinedProcessor):
    """
    Extended CombinedProcessor that integrates global usage data.
    """
    
    def __init__(self, delivery_file: str, inventory_file: str, global_usage_file: str):
        """
        Initialize the Global Usage processor.
        
        Args:
            delivery_file (str): Path to the delivery Excel file
            inventory_file (str): Path to the inventory Excel file
            global_usage_file (str): Path to the global usage CSV file
        """
        super().__init__(delivery_file, inventory_file)
        self.global_usage_file = global_usage_file
        self.global_usage_data = {}
        self.ingredient_mapping = {}
        
    def _load_global_usage_data(self) -> Dict[str, pd.DataFrame]:
        """
        Load global usage data from CSV file.
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing the global usage data
        """
        try:
            # Read the CSV file
            df = pd.read_csv(self.global_usage_file, encoding='utf-8')
            logger.info(f"Loaded global usage data with {len(df)} rows")
            
            # Clean column names (remove newlines and extra spaces)
            df.columns = [col.replace('\n', ' ').strip() for col in df.columns]
            
            # Clean the data
            df = self._clean_global_usage_data(df)
            
            # Store the processed data
            self.global_usage_data['global_usage'] = df
            
            logger.info(f"Processed global usage data: {df.shape}")
            logger.info(f"Columns: {list(df.columns)}")
            
            return self.global_usage_data
            
        except Exception as e:
            logger.error(f"Error loading global usage data: {e}")
            return {}
    
    def _clean_global_usage_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean and process the global usage data.
        
        Args:
            df (pd.DataFrame): Raw global usage data
            
        Returns:
            pd.DataFrame: Cleaned global usage data
        """
        if df.empty:
            return df
        
        df_clean = df.copy()
        
        # Clean ingredient names - handle the actual column name with extra spaces
        ingredient_name_col = 'Ingredient  Name'  # Note the double space
        if ingredient_name_col in df_clean.columns:
            df_clean['Ingredient Name'] = df_clean[ingredient_name_col].str.strip()
            # Remove rows with empty ingredient names
            df_clean = df_clean[df_clean['Ingredient Name'].notna() & (df_clean['Ingredient Name'] != '')]
        
        # Clean ingredient amounts - convert to numeric, handling commas
        if 'Ingredient Amount' in df_clean.columns:
            df_clean['Ingredient Amount'] = df_clean['Ingredient Amount'].astype(str)
            df_clean['Ingredient Amount'] = df_clean['Ingredient Amount'].str.replace(',', '')
            df_clean['Ingredient Amount'] = pd.to_numeric(df_clean['Ingredient Amount'], errors='coerce')
            # Fill NaN with 0
            df_clean['Ingredient Amount'] = df_clean['Ingredient Amount'].fillna(0)
        
        # Clean ingredient stock - convert to numeric, handling commas
        if 'Ingredient Stock' in df_clean.columns:
            df_clean['Ingredient Stock'] = df_clean['Ingredient Stock'].astype(str)
            df_clean['Ingredient Stock'] = df_clean['Ingredient Stock'].str.replace(',', '')
            df_clean['Ingredient Stock'] = pd.to_numeric(df_clean['Ingredient Stock'], errors='coerce')
            # Fill NaN with 0
            df_clean['Ingredient Stock'] = df_clean['Ingredient Stock'].fillna(0)
        
        logger.info(f"Cleaned global usage data: {df_clean.shape}")
        return df_clean
    
    def _create_ingredient_mapping(self) -> Dict[str, str]:
        """
        Create a mapping between ingredient names and product names from product categories.
        New ingredients will be added to product_categories.csv for manual categorization later.
        
        Returns:
            Dict[str, str]: Mapping from ingredient names to product names
        """
        try:
            # Load product categories
            product_categories = self._load_product_categories()
            
            # Create ingredient mapping
            ingredient_mapping = {}
            new_ingredients = []
            
            # Load global usage data if not already loaded
            if not self.global_usage_data:
                self._load_global_usage_data()
            
            if 'global_usage' in self.global_usage_data:
                global_df = self.global_usage_data['global_usage']
                
                for _, row in global_df.iterrows():
                    ingredient_name = row.get('Ingredient Name', '')
                    if pd.notna(ingredient_name) and ingredient_name != '':
                        # Try to find a matching product in the categories
                        # Look for exact matches first
                        if ingredient_name in product_categories:
                            ingredient_mapping[ingredient_name] = ingredient_name
                        else:
                            # Try partial matches
                            matched_product = None
                            for product in product_categories.keys():
                                if (ingredient_name.lower() in product.lower() or 
                                    product.lower() in ingredient_name.lower()):
                                    matched_product = product
                                    break
                            
                            if matched_product:
                                ingredient_mapping[ingredient_name] = matched_product
                            else:
                                # No match found, treat ingredient as a new product
                                ingredient_mapping[ingredient_name] = ingredient_name
                                new_ingredients.append(ingredient_name)
            
            self.ingredient_mapping = ingredient_mapping
            
            # Add new ingredients to product_categories.csv for manual categorization
            if new_ingredients:
                self._add_new_ingredients_to_categories(new_ingredients)
                logger.info(f"Added {len(new_ingredients)} new ingredients to product_categories.csv for manual categorization")
            
            logger.info(f"Created ingredient mapping with {len(ingredient_mapping)} entries")
            logger.info(f"New ingredients added: {', '.join(new_ingredients)}")
            
            return ingredient_mapping
            
        except Exception as e:
            logger.error(f"Error creating ingredient mapping: {e}")
            return {}
    
    def _add_new_ingredients_to_categories(self, new_ingredients: List[str]):
        """
        Add new ingredients to product_categories.csv with 'N/A' category for manual editing later.
        
        Args:
            new_ingredients (List[str]): List of new ingredient names to add
        """
        try:
            # Load existing categories
            existing_categories = self._load_product_categories()
            
            # Add new ingredients with 'N/A' category
            for ingredient in new_ingredients:
                if ingredient not in existing_categories:
                    existing_categories[ingredient] = 'N/A'
            
            # Save updated categories
            self._save_product_categories(existing_categories)
            
            logger.info(f"Added {len(new_ingredients)} new ingredients to product categories")
            
        except Exception as e:
            logger.error(f"Error adding new ingredients to categories: {e}")
    
    def _calculate_monthly_usage(self, ingredient_name: str) -> float:
        """
        Calculate monthly usage for a given ingredient.
        
        Args:
            ingredient_name (str): Name of the ingredient
            
        Returns:
            float: Monthly usage amount
        """
        try:
            if 'global_usage' not in self.global_usage_data:
                return 0.0
            
            global_df = self.global_usage_data['global_usage']
            
            # Find the ingredient in global usage data
            ingredient_row = global_df[global_df['Ingredient Name'] == ingredient_name]
            
            if not ingredient_row.empty:
                amount = ingredient_row.iloc[0].get('Ingredient Amount', 0)
                if pd.notna(amount):
                    # Convert to monthly usage (assuming the amount is annual)
                    monthly_usage = amount / 12.0
                    return monthly_usage
            
            return 0.0
            
        except Exception as e:
            logger.error(f"Error calculating monthly usage for {ingredient_name}: {e}")
            return 0.0
    
    def _calculate_projected_inventory(self, total_inventory: float, monthly_usage: float, months: int) -> float:
        """
        Calculate projected inventory for a given number of months.
        
        Args:
            total_inventory (float): Total current inventory
            monthly_usage (float): Monthly usage amount
            months (int): Number of months to project
            
        Returns:
            float: Projected inventory after specified months
        """
        try:
            projected = total_inventory - (monthly_usage * months)
            return projected
        except Exception as e:
            logger.error(f"Error calculating projected inventory: {e}")
            return total_inventory
    
    def create_enhanced_pivot_table(self, combined_df: pd.DataFrame = None) -> pd.DataFrame:
        """
        Create an enhanced pivot table with monthly usage and projected inventory.
        
        Args:
            combined_df (pd.DataFrame, optional): Combined dataframe. If None, will use self.combine_all_data()
            
        Returns:
            pd.DataFrame: Enhanced pivot table with usage and projections
        """
        if combined_df is None:
            combined_df = self.combine_all_data()
        
        if combined_df.empty:
            logger.warning("No data available for enhanced pivot table")
            return pd.DataFrame()
        
        # Create the base pivot table
        base_pivot = self.create_pivot_table(combined_df)
        
        if base_pivot.empty:
            logger.warning("Failed to create base pivot table")
            return pd.DataFrame()
        
        # Load global usage data and create ingredient mapping
        self._load_global_usage_data()
        self._create_ingredient_mapping()
        
        # Create enhanced pivot table
        enhanced_pivot = base_pivot.copy()
        
        # Add Monthly Usage column
        enhanced_pivot['Monthly Usage'] = 0.0
        
        # Add projected inventory columns
        for month in range(1, 7):
            enhanced_pivot[f'{month} Month Projection'] = 0.0
        
        # Calculate monthly usage and projections for each category
        for category in enhanced_pivot.index:
            if category == 'TOTAL':
                continue
                
            # Get the total inventory for this category
            category_total = enhanced_pivot.loc[category, 'TOTAL']
            
            # Calculate monthly usage for this category
            monthly_usage = 0.0
            
            # Find products in this category and sum their usage
            category_products = combined_df[combined_df['CATEGORY'] == category]['PRODUCT'].unique()
            
            for product in category_products:
                # Find matching ingredient in global usage
                for ingredient_name, mapped_product in self.ingredient_mapping.items():
                    if mapped_product == product:
                        product_usage = self._calculate_monthly_usage(ingredient_name)
                        monthly_usage += product_usage
                        break
                else:
                    # If no mapping found, try to find the product directly in global usage
                    # This handles the case where ingredients are treated as products
                    direct_usage = self._calculate_monthly_usage(product)
                    monthly_usage += direct_usage
            
            # Set monthly usage
            enhanced_pivot.loc[category, 'Monthly Usage'] = monthly_usage
            
            # Calculate projections
            for month in range(1, 7):
                projection = self._calculate_projected_inventory(category_total, monthly_usage, month)
                enhanced_pivot.loc[category, f'{month} Month Projection'] = projection
        
        # Calculate totals for new columns
        enhanced_pivot.loc['TOTAL', 'Monthly Usage'] = enhanced_pivot['Monthly Usage'].sum()
        
        for month in range(1, 7):
            col_name = f'{month} Month Projection'
            enhanced_pivot.loc['TOTAL', col_name] = enhanced_pivot[col_name].sum()
        
        logger.info(f"Created enhanced pivot table with shape: {enhanced_pivot.shape}")
        return enhanced_pivot
    
    def save_enhanced_pivot_table(self, enhanced_pivot_df: pd.DataFrame, output_dir: str = None):
        """
        Save the enhanced pivot table to CSV and Excel files with formatting.
        
        Args:
            enhanced_pivot_df (pd.DataFrame): Enhanced pivot table to save
            output_dir (str): Directory to save the pivot table. If None, uses self.output_dir
        """
        if enhanced_pivot_df.empty:
            logger.warning("No enhanced pivot table to save")
            return
        
        # Use instance output directory if none specified
        if output_dir is None:
            output_dir = self.output_dir
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"Saving enhanced pivot table to: {output_dir}")
        
        # Save to CSV
        csv_file = os.path.join(output_dir, "ENHANCED_PIVOT_TABLE.csv")
        enhanced_pivot_df.to_csv(csv_file)
        logger.info(f"Saved enhanced pivot table to {csv_file}")
        
        # Save to Excel with enhanced formatting
        excel_file = os.path.join(output_dir, "ENHANCED_PIVOT_TABLE.xlsx")
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            enhanced_pivot_df.to_excel(writer, sheet_name='Enhanced Pivot Table')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Enhanced Pivot Table']
            
            # Format the worksheet
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            usage_font = Font(bold=True, color="000000")
            usage_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            projection_font = Font(bold=True, color="000000")
            projection_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            negative_font = Font(bold=True, color="FF0000")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Format row headers (categories)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.border = border
                
                # Highlight total row
                if cell.value == 'TOTAL':
                    for col in range(1, worksheet.max_column + 1):
                        total_cell = worksheet.cell(row=row, column=col)
                        total_cell.font = total_font
                        total_cell.fill = total_fill
                        total_cell.border = border
            
            # Format all data cells
            for row in range(2, worksheet.max_row + 1):
                for col in range(2, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
                    
                    # Get column header to determine formatting
                    col_header = worksheet.cell(row=1, column=col).value
                    
                    # Format Monthly Usage column
                    if col_header == 'Monthly Usage':
                        cell.fill = usage_fill
                        cell.font = usage_font
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        else:
                            cell.number_format = '#,##0.00'
                    
                    # Format projection columns
                    elif 'Month Projection' in str(col_header):
                        cell.fill = projection_fill
                        cell.font = projection_font
                        
                        # Check if value is negative and color red
                        if isinstance(cell.value, (int, float)) and cell.value < 0:
                            cell.font = negative_font
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        else:
                            cell.number_format = '#,##0.00'
                    
                    # Format regular data cells
                    else:
                        if cell.value == 0 or cell.value == 0.0:
                            cell.value = '-'
                            cell.font = Font(color="808080")
                        elif isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell.number_format = '#,##0'
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Saved enhanced formatted pivot table to {excel_file}")
    
    def process_all_data(self) -> Dict[str, pd.DataFrame]:
        """
        Process all data including global usage.
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing all processed data
        """
        # Process base data using parent class
        base_results = super().process_all_data()
        
        # Process global usage data
        global_usage_results = self._load_global_usage_data()
        
        # Combine results
        all_results = {**base_results, **global_usage_results}
        
        logger.info(f"Processed all data: {len(all_results)} datasets")
        return all_results


def main():
    """
    Main function to demonstrate the Global Usage processor usage.
    """
    import os
    
    # Initialize the processor
    processor = GlobalUsageProcessor(
        'input/deliveries.xlsx', 
        'input/inventory.xlsx',
        'input/global_usage.csv'
    )
    
    # Process all data
    results = processor.process_all_data()
    
    # Print summary
    summary = processor.get_summary()
    print("\n=== GLOBAL USAGE PROCESSING SUMMARY ===")
    for sheet_name, info in summary.items():
        print(f"\n{sheet_name.upper()}:")
        print(f"  Rows: {info['rows']}")
        print(f"  Columns: {info['columns']}")
        if info['column_names']:
            print(f"  Columns: {', '.join(info['column_names'])}")
    
    # Save processed data
    processor.save_processed_data()
    
    # Combine all data
    combined_df = processor.combine_all_data()
    
    # Save combined data
    if not combined_df.empty:
        combined_file = os.path.join(processor.output_dir, "COMBINED_ALL_DATA.csv")
        combined_df.to_csv(combined_file, index=False)
        logger.info(f"Saved combined data to {combined_file}")
        print(f"\n=== COMBINED DATA SUMMARY ===")
        print(f"Total rows: {len(combined_df)}")
        print(f"Status breakdown: {combined_df['STATUS'].value_counts().to_dict()}")
        print(f"Category breakdown: {combined_df['CATEGORY'].value_counts().to_dict()}")
        
        # Create and save enhanced pivot table
        print(f"\n=== CREATING ENHANCED PIVOT TABLE ===")
        
        enhanced_pivot_table = processor.create_enhanced_pivot_table(combined_df)
        
        if not enhanced_pivot_table.empty:
            processor.save_enhanced_pivot_table(enhanced_pivot_table)
            print(f"✅ Enhanced pivot table created successfully!")
            print(f"Shape: {enhanced_pivot_table.shape}")
            print(f"Categories: {len(enhanced_pivot_table.index) - 1}")  # Exclude TOTAL row
            print(f"Columns: {len(enhanced_pivot_table.columns)}")
            print(f"\nEnhanced Pivot Table Preview:")
            print(enhanced_pivot_table.to_string())
        else:
            print("❌ Failed to create enhanced pivot table")
    
    print(f"\n=== INDIVIDUAL DATA SAMPLES ===")
    for sheet_name, df in results.items():
        if not df.empty:
            print(f"\n{sheet_name}:")
            print(f"Shape: {df.shape}")
            print("First 5 rows:")
            print(df.head().to_string(index=False))
            print("\n" + "="*50)


if __name__ == "__main__":
    main()
