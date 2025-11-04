#!/usr/bin/env python3
import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('input/20251104/local.xlsx', data_only=True)
ws = wb['UNSERVED LOCAL']

print('=== Excel Structure ===')
print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}')

# Find header row
header_row = None
for i in range(1, min(20, ws.max_row+1)):
    val = ws.cell(row=i, column=1).value
    if val and str(val).strip().upper() == 'RAW MATERIALS':
        header_row = i
        print(f'\nFound RAW MATERIALS header at row {i}')
        break

if header_row:
    # Find RAW MATERIALS column
    raw_materials_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip().upper() == 'RAW MATERIALS':
            raw_materials_col = col
            print(f'RAW MATERIALS is in column {col}')
            break
    
    if raw_materials_col:
        print(f'\n=== Checking US SOYA entries (rows {header_row+1} to {min(header_row+30, ws.max_row+1)}) ===')
        us_soya_count = 0
        for row in range(header_row + 1, min(header_row + 30, ws.max_row + 1)):
            val = ws.cell(row=row, column=raw_materials_col).value
            if val and 'US SOYA' in str(val).upper():
                us_soya_count += 1
                # Check if merged
                cell = ws.cell(row=row, column=raw_materials_col)
                merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        merged = True
                        print(f'Row {row}: US SOYA (MERGED - spans rows {merged_range.min_row} to {merged_range.max_row})')
                        # Get other columns
                        supplier = ws.cell(row=row, column=2).value if ws.max_column >= 2 else None
                        qty_col = None
                        for c in range(1, ws.max_column + 1):
                            hdr = ws.cell(row=header_row, column=c).value
                            if hdr and 'DELIVER' in str(hdr).upper() and 'KILO' in str(hdr).upper():
                                qty_col = c
                                break
                        qty = ws.cell(row=row, column=qty_col).value if qty_col else None
                        print(f'  Supplier: {supplier}, Qty: {qty}')
                        break
                if not merged:
                    supplier = ws.cell(row=row, column=2).value if ws.max_column >= 2 else None
                    print(f'Row {row}: US SOYA (NOT MERGED) - Supplier: {supplier}')
        
        print(f'\nTotal US SOYA entries found: {us_soya_count}')

