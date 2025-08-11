#!/usr/bin/env python3
"""
Inventory Excel Processor

A specialized processor for handling inventory.xlsx files.
"""

import pandas as pd
import numpy as np
import json
from typing import Dict, List, Tuple, Optional
import logging
import os
from datetime import datetime
import openpyxl

# Set up logging
os.makedirs('logs', exist_ok=True)

# Create a unique log filename with timestamp
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
log_filename = f'logs/inventory_processor_{timestamp}.log'

# Set up logging to both file and console
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, mode='w'),
        logging.StreamHandler()  # This keeps console output
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"Logging to file: {log_filename}")

class InventoryProcessor:
    """
    A class to process inventory Excel files.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the Inventory processor.
        
        Args:
            file_path (str): Path to the inventory Excel file
        """
        self.file_path = file_path
        self.excel_file = None
        self.sheets_data = {}
        
    def load_excel_file(self) -> bool:
        """
        Load the Excel file and get sheet information.
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            self.excel_file = pd.ExcelFile(self.file_path)
            logger.info(f"Successfully loaded Excel file: {self.file_path}")
            logger.info(f"Available sheets: {self.excel_file.sheet_names}")
            return True
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return False
    
    def _extract_inventory_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Extract inventory data from the complex structure.
        
        Args:
            df (pd.DataFrame): Raw dataframe from Excel
            
        Returns:
            pd.DataFrame: Cleaned inventory data
        """
        if df.empty:
            return df
        
        # Find the header row (look for "ITEM DESCRIPTION" and "QTY IN KGS")
        header_row = -1
        for i in range(min(25, len(df))):
            row_values = [str(val).strip() if pd.notna(val) else '' for val in df.iloc[i]]
            if 'ITEM DESCRIPTION' in row_values and 'QTY IN KGS' in row_values:
                header_row = i
                break
        
        if header_row == -1:
            logger.warning("Could not find inventory header row")
            return df
        
        logger.info(f"Found header row at index {header_row}")
        
        # Extract headers
        headers = df.iloc[header_row]
        clean_headers = []
        for i, header in enumerate(headers):
            if pd.notna(header) and str(header).strip() != '':
                clean_headers.append(str(header).strip())
            else:
                clean_headers.append(f'Column_{i}')
        
        # Start data from the row after header
        data_df = df.iloc[header_row + 1:].copy()
        data_df.columns = clean_headers
        
        # Clean the data
        data_df = data_df.dropna(how='all')  # Remove completely empty rows
        
        # Extract only the relevant columns (ITEM DESCRIPTION and QTY IN KGS)
        relevant_columns = []
        for col in data_df.columns:
            if 'ITEM DESCRIPTION' in col or 'QTY IN KGS' in col or 'DESCRIPTION' in col or 'QTY' in col:
                relevant_columns.append(col)
        
        if relevant_columns:
            data_df = data_df[relevant_columns]
        
        # --- openpyxl merged cell fill for ITEM DESCRIPTION ---
        if 'ITEM DESCRIPTION' in data_df.columns:
            try:
                wb = openpyxl.load_workbook(self.file_path, data_only=True)
                ws = wb[wb.sheetnames[0]]
                merged_ranges = ws.merged_cells.ranges
                # Find header row in openpyxl
                header_row_idx = None
                for i in range(1, 26):
                    if ws.cell(row=i, column=1).value and str(ws.cell(row=i, column=1).value).strip().upper() == 'ITEM DESCRIPTION':
                        header_row_idx = i
                        break
                if header_row_idx is not None:
                    data_start_row = header_row_idx + 1
                    excel_row_for_df = lambda df_idx: data_start_row + df_idx
                    filled_values = []
                    for df_idx, value in enumerate(data_df['ITEM DESCRIPTION']):
                        excel_row = excel_row_for_df(df_idx)
                        cell = ws.cell(row=excel_row, column=1)
                        is_merged = any([cell.coordinate in rng for rng in merged_ranges])
                        if is_merged:
                            for rng in merged_ranges:
                                if cell.coordinate in rng:
                                    top_left = ws.cell(row=rng.min_row, column=rng.min_col)
                                    filled_values.append(top_left.value)
                                    break
                        else:
                            filled_values.append(value)
                    data_df['ITEM DESCRIPTION'] = filled_values
                    logger.info("Applied openpyxl-based merged cell filling for ITEM DESCRIPTION")
            except Exception as e:
                logger.warning(f'openpyxl merged cell handling failed in inventory: {e}')
        # --- end openpyxl merged cell fill ---
        
        # Clean up the data
        for col in data_df.columns:
            if data_df[col].dtype == 'object':
                data_df[col] = data_df[col].fillna('')
        
        # Remove rows that are completely empty or contain only formatting
        data_df = data_df[data_df.apply(lambda row: row.astype(str).str.strip().ne('').any(), axis=1)]
        
        logger.info(f"Extracted inventory data: {len(data_df)} rows, {len(data_df.columns)} columns")
        return data_df
    
    def _clean_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean the table by removing empty rows and columns.
        
        Args:
            df (pd.DataFrame): The dataframe to clean
            
        Returns:
            pd.DataFrame: Cleaned dataframe
        """
        if df.empty:
            return df
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Fill NaN values with empty strings for string columns
        for col in df.columns:
            if hasattr(df[col], 'dtype'):
                if df[col].dtype == 'object':
                    df[col] = df[col].fillna('')
        
        return df
    
    def process_inventory_sheet(self, sheet_name: str = None) -> pd.DataFrame:
        """
        Process the inventory sheet.
        
        Args:
            sheet_name (str): Name of the sheet to process. If None, uses the first sheet.
            
        Returns:
            pd.DataFrame: Processed inventory data
        """
        try:
            if sheet_name is None:
                sheet_name = self.excel_file.sheet_names[0]
            
            logger.info(f"Processing {sheet_name} sheet")
            
            # Read the sheet
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            logger.info(f"Loaded {sheet_name} sheet with {len(df)} rows")
            
            # Extract inventory data from the complex structure
            df = self._extract_inventory_data(df)
            
            # Clean the table
            df = self._clean_table(df)
            
            logger.info(f"Processed {sheet_name} sheet: {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            logger.error(f"Error processing {sheet_name} sheet: {e}")
            return pd.DataFrame()
    
    def process_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """
        Process all sheets in the inventory file.
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary of processed dataframes
        """
        if not self.load_excel_file():
            return {}
        
        results = {}
        
        for sheet_name in self.excel_file.sheet_names:
            try:
                df = self.process_inventory_sheet(sheet_name)
                if not df.empty:
                    results[sheet_name] = df
                    logger.info(f"Successfully processed {sheet_name} sheet")
                else:
                    logger.warning(f"Empty result for {sheet_name} sheet")
            except Exception as e:
                logger.error(f"Error processing {sheet_name} sheet: {e}")
        
        self.sheets_data = results
        return results
    
    def save_processed_data(self, output_dir: str = "processed_data"):
        """
        Save all processed data to CSV files.
        
        Args:
            output_dir (str): Directory to save the processed data
        """
        import os
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        for sheet_name, df in self.sheets_data.items():
            if not df.empty:
                output_file = os.path.join(output_dir, "inventory.csv")
                df.to_csv(output_file, index=False)
                logger.info(f"Saved {sheet_name} to {output_file}")
    
    def get_summary(self) -> Dict[str, dict]:
        """
        Get a summary of all processed data.
        
        Returns:
            Dict[str, dict]: Summary information for each sheet
        """
        summary = {}
        
        for sheet_name, df in self.sheets_data.items():
            if not df.empty:
                summary[sheet_name] = {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': list(df.columns),
                    'data_types': df.dtypes.to_dict(),
                    'missing_values': df.isnull().sum().to_dict()
                }
            else:
                summary[sheet_name] = {
                    'rows': 0,
                    'columns': 0,
                    'column_names': [],
                    'data_types': {},
                    'missing_values': {}
                }
        
        return summary


def main():
    """
    Main function to demonstrate the Inventory processor usage.
    """
    import os
    
    # Initialize the processor
    processor = InventoryProcessor('input/inventory.xlsx')
    
    # Process all sheets
    results = processor.process_all_sheets()
    
    # Print summary
    summary = processor.get_summary()
    print("\n=== INVENTORY PROCESSING SUMMARY ===")
    for sheet_name, info in summary.items():
        print(f"\n{sheet_name.upper()}:")
        print(f"  Rows: {info['rows']}")
        print(f"  Columns: {info['columns']}")
        if info['column_names']:
            print(f"  Columns: {', '.join(info['column_names'])}")
    
    # Save processed data
    processor.save_processed_data()
    
    print(f"\n=== INVENTORY DATA SAMPLE ===")
    for sheet_name, df in results.items():
        if not df.empty:
            print(f"\n{sheet_name}:")
            print(f"Shape: {df.shape}")
            print("First 10 rows:")
            print(df.head(10).to_string(index=False))
            print("\n" + "="*50)


if __name__ == "__main__":
    main() 